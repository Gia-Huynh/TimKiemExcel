from openpyxl import load_workbook

def row_to_string(row):
    return ', '.join(str(cell) for cell in row)

def get_rows_as_strings(file_path):
    wb = load_workbook(file_path)
    ws = wb.active
    row_strings = []
    for row in ws.iter_rows(values_only=True):
        row_strings.append(row_to_string(row))
    return row_strings

# Example usage
file_path = './Data/GUI DIEU 6.2024.xlsx'
row_strings = get_rows_as_strings(file_path)
for row_string in row_strings:
    print(row_string)
